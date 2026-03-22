"""
langgraph_orchestrator.py
=========================
ETL — 10 Agents wired as a LangGraph StateGraph.

Architecture:
  ┌─────────────────────────────────────────────────────────────────┐
  │                    LANGGRAPH STATE MACHINE                       │
  │                                                                   │
  │  START → schema_analyzer → rule_interpreter → precision_guard    │
  │       → db_mapping → output_mapping → drift_detection            │
  │       → [conditional: drift?] → rule_optimizer                   │
  │       → test_generator → anomaly_detector → excel_writer         │
  │       → runtime_validator → [conditional: pass?] → END           │
  │                              ↓ FAIL                               │
  │                         rule_optimizer (re-run)                   │
  └─────────────────────────────────────────────────────────────────┘

Each node:
  - Receives full ETLState dict
  - Returns a partial dict (only changed keys)
  - StateGraph merges updates automatically

Supports: Claude | Gemini | Ollama   (via create_llm factory)
CLI:      python langgraph_orchestrator.py --provider gemini --input data.xlsx --rules rules.xlsx
"""

from __future__ import annotations
import argparse, json, os, sys, time
from datetime import datetime, timezone
from pathlib import Path
from decimal import Decimal
from typing import TypedDict, Annotated, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── LangChain / LangGraph (native impl) ──────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from lc_core import (
    ChatPromptTemplate, JsonOutputParser, StrOutputParser,
    RunnableLambda, RunnablePassthrough, create_llm, LCChain,
    HumanMessage, SystemMessage, AIMessage,
)
from lg_graph import (
    StateGraph, MemorySaver, END, AgentState,
)

# ── ETL engine ────────────────────────────────────────────────────────────────
from arithmetic_engine import (
    compute_multiply, compute_subtract, compute_divide, compute_add_constant, to_decimal,
)
from mock_db import db_lookup


# ══════════════════════════════════════════════════════════════════════════════
# GRAPH STATE
# ══════════════════════════════════════════════════════════════════════════════

class ETLState(AgentState):
    """
    Full graph state. Every key is optional — nodes update only what they touch.

    LangGraph merges partial updates: if a node returns {"schema": x},
    only state["schema"] is updated; everything else is unchanged.
    """
    pass


def initial_state(input_file: str, rule_file: str, out_path: str,
                  provider: str, model: str) -> ETLState:
    return ETLState({
        # ── Config ──────────────────────────────────────────────────────────
        "input_file":      input_file,
        "rule_file":       rule_file,
        "out_path":        out_path,
        "provider":        provider,
        "model":           model,
        "timestamp":       datetime.now(timezone.utc).isoformat(),
        # ── Data ────────────────────────────────────────────────────────────
        "input_columns":   [],
        "input_rows":      [],
        "output_columns":  [],
        "raw_rules":       [],
        # ── Agent outputs ────────────────────────────────────────────────────
        "schema":          {},
        "interpreted_rules": [],
        "db_resolved":     {},
        "output_rows":     [],
        "drift_result":    {},
        "opt_result":      {},
        "test_cases":      [],
        "test_results":    {},
        "anomaly_result":  {},
        "expected_rows":   [],
        "spec_map":        {},
        "validation":      {},
        # ── Control flow ─────────────────────────────────────────────────────
        "drift_detected":  False,
        "validation_pass": False,
        "retry_count":     0,
        "max_retries":     2,
        "agent_log":       [],
        "errors":          [],
    })


# ══════════════════════════════════════════════════════════════════════════════
# LCEL CHAIN BUILDERS — each agent builds prompt | llm | parser chains
# ══════════════════════════════════════════════════════════════════════════════

def _make_schema_chain(llm) -> LCChain:
    prompt = ChatPromptTemplate.from_messages([
        ("system", """You are a financial data schema expert.
Analyze column names and sample data from ANY financial Excel file.
For each column return semantic_type, data_type, nullable, description.
Return ONLY JSON: {{"columns":[{{"name":"...","semantic_type":"...","data_type":"...","nullable":true,"description":"..."}}]}}"""),
        ("human", "Columns: {columns}\nSample rows: {sample_rows}"),
    ])
    return prompt | llm | JsonOutputParser()


def _make_rule_chain(llm) -> LCChain:
    prompt = ChatPromptTemplate.from_messages([
        ("system", """You are an ETL rule interpreter for financial data.
Available input columns: {{input_columns}}
Convert each English business rule to a structured operation.
Operations: copy | multiply | subtract | divide | add_constant | db_lookup
Return ONLY a JSON array — one object per rule:
[{{"output_col":"...","operation":"...","operand_a":"...","operand_b":"...","constant":null,"db_key":null,"reasoning":"..."}}]"""),
        ("human", "Input columns: {input_columns}\nRules:\n{rules_json}"),
    ])
    return prompt | llm | JsonOutputParser()


def _make_drift_chain(llm) -> LCChain:
    prompt = ChatPromptTemplate.from_messages([
        ("system", """You are a data drift detection specialist.
Check input→output samples for systematic errors.
Return JSON: {{"drift_detected":false,"issues":[{{"field":"...","issue":"...","severity":"high|medium|low"}}],"summary":"..."}}"""),
        ("human", "Rules (first 15):\n{rules_sample}\n\nInput→Output samples:\n{io_samples}"),
    ])
    return prompt | llm | JsonOutputParser()


def _make_opt_chain(llm) -> LCChain:
    prompt = ChatPromptTemplate.from_messages([
        ("system", """ETL optimization expert. Suggest batching, caching, de-duplication.
Return JSON: {{"optimizations":[{{"type":"...","description":"...","impact":"high|medium|low","affected_fields":[]}}],"summary":"..."}}"""),
        ("human", "Operation distribution: {op_dist}\nSample rules: {rules_sample}"),
    ])
    return prompt | llm | JsonOutputParser()


def _make_test_chain(llm) -> LCChain:
    prompt = ChatPromptTemplate.from_messages([
        ("system", """You are a financial ETL QA engineer.
For each rule create a happy-path test AND an edge-case test.
Return JSON array: [{{"rule_field":"...","operation":"...","test_name":"...","inputs":{{}},"expected_formula":"...","edge_case":false}}]"""),
        ("human", "Input columns: {input_columns}\nRules (first 20):\n{rules_sample}\nSample input row:\n{sample_row}"),
    ])
    return prompt | llm | JsonOutputParser()


def _make_anomaly_chain(llm) -> LCChain:
    prompt = ChatPromptTemplate.from_messages([
        ("system", """Financial data anomaly detection expert.
Flag: all-zero fields, extreme outliers (>10x mean), no-variation fields.
Return JSON: {{"anomalies":[{{"field":"...","type":"...","detail":"...","severity":"high|medium|low"}}],"clean_fields":0,"summary":"..."}}"""),
        ("human", "Field statistics:\n{stats}\nRows: {row_count}"),
    ])
    return prompt | llm | JsonOutputParser()


def _make_spec_chain(llm) -> LCChain:
    prompt = ChatPromptTemplate.from_messages([
        ("system", """QA engineer writing test specifications.
For each rule write a clear validation spec.
Return JSON array: [{{"field":"...","spec":"...","category":"arithmetic|copy|db_lookup|other"}}]"""),
        ("human", "Rules:\n{rules_sample}\nSample expected (row 1):\n{sample_exp}"),
    ])
    return prompt | llm | JsonOutputParser()


# ══════════════════════════════════════════════════════════════════════════════
# DETERMINISTIC HELPERS (no LLM)
# ══════════════════════════════════════════════════════════════════════════════

def _read_rule_mapping(path: str):
    import re
    df = pd.read_excel(path, header=None, dtype=str).fillna("")
    header_row = 0
    for i, row in df.iterrows():
        if sum(1 for v in row if v and not _is_num(str(v))) >= 2:
            header_row = i; break
    df = pd.read_excel(path, header=header_row, dtype=str).fillna("")
    df.columns = ["output_col","business_rule","input_columns"] + list(df.columns[3:])
    df = df[df["output_col"].str.strip() != ""]
    return df["output_col"].tolist(), df[["output_col","business_rule","input_columns"]].to_dict("records")

def _is_num(s):
    try: Decimal(s); return True
    except: return False

def _regex_interpret(rules, input_columns):
    import re
    out = []
    for r in rules:
        txt = r.get("business_rule","").lower()
        oc  = r.get("output_col","")
        ic  = [c.strip() for c in str(r.get("input_columns","")).split(",")]
        m_db  = re.search(r"key\s+(key_\d+)", txt, re.I)
        m_con = re.search(r"constant value\s+(\d+(?:\.\d+)?)", txt)
        if m_db:
            op = {"operation":"db_lookup","operand_a":None,"operand_b":None,"constant":None,"db_key":m_db.group(1).upper()}
        elif "multiply" in txt:
            op = {"operation":"multiply","operand_a":ic[0] if ic else None,"operand_b":ic[1] if len(ic)>1 else None,"constant":None,"db_key":None}
        elif m_con and "add" in txt:
            op = {"operation":"add_constant","operand_a":ic[0] if ic else None,"operand_b":None,"constant":float(m_con.group(1)),"db_key":None}
        elif "subtract" in txt or "net price" in txt:
            op = {"operation":"subtract","operand_a":ic[0] if ic else None,"operand_b":ic[1] if len(ic)>1 else None,"constant":None,"db_key":None}
        elif "divide" in txt:
            op = {"operation":"divide","operand_a":ic[0] if ic else None,"operand_b":ic[1] if len(ic)>1 else None,"constant":None,"db_key":None}
        else:
            op = {"operation":"copy","operand_a":ic[0] if ic else None,"operand_b":None,"constant":None,"db_key":None}
        out.append({"output_col":oc,"reasoning":"regex-fallback",**op})
    return out

def _execute_row(rules, row, ri, db_resolved):
    out = {}
    for rule in rules:
        col, op = rule.get("output_col",""), rule.get("operation","")
        oa, ob, const = rule.get("operand_a"), rule.get("operand_b"), rule.get("constant")
        va = row.get(oa,"") if oa else ""
        vb = row.get(ob,"") if ob else ""
        if   op=="copy":          out[col]=str(row.get(oa,"")).strip()
        elif op=="multiply":      out[col]=compute_multiply(va,vb)
        elif op=="subtract":      out[col]=compute_subtract(va,vb)
        elif op=="divide":        out[col]=compute_divide(va,vb)
        elif op=="add_constant":  out[col]=compute_add_constant(va,const)
        elif op=="db_lookup":     out[col]=db_resolved.get((ri,col),"N/A")
        else:                     out[col]=""
    return out

def _compare_cell(actual, expected, op):
    if str(actual)==str(expected): return "PASS",""
    try:
        diff=abs(to_decimal(actual)-to_decimal(expected))
        if diff<=Decimal("0.00005"): return "PASS",""
        return "FAIL",f"expected={expected} actual={actual} diff={diff}"
    except: pass
    return ("PASS","") if str(actual).strip()==str(expected).strip() else ("FAIL",f"expected='{expected}' actual='{actual}'")


# ══════════════════════════════════════════════════════════════════════════════
# GRAPH NODE FUNCTIONS
# Each receives state: ETLState, returns partial dict of changes
# ══════════════════════════════════════════════════════════════════════════════

def node_loader(state: ETLState) -> dict:
    """Load input Excel + rule mapping. Pure deterministic."""
    _log(state, "loader", "Loading input files...")
    out_cols, raw_rules = _read_rule_mapping(state["rule_file"])
    df = pd.read_excel(state["input_file"], dtype=str).fillna("")
    input_columns = list(df.columns)
    input_rows    = df.to_dict("records")
    _log(state, "loader", f"✅ {len(raw_rules)} rules, {len(input_rows)} data rows")
    return {"output_columns": out_cols, "raw_rules": raw_rules,
            "input_columns": input_columns, "input_rows": input_rows}


def node_schema_analyzer(state: ETLState) -> dict:
    """Agent 1 — LCEL chain: prompt | llm | JsonOutputParser."""
    _log(state, "schema_analyzer", f"Analyzing {len(state['input_columns'])} columns...")
    llm   = _get_llm(state)
    chain = _make_schema_chain(llm)
    result = chain.invoke({
        "columns":     str(state["input_columns"]),
        "sample_rows": json.dumps(state["input_rows"][:3], default=str),
    })
    if not result or "columns" not in result:
        result = {"columns": [{"name": c, "data_type": "string", "semantic_type": "unknown",
                                "nullable": True, "description": c}
                               for c in state["input_columns"]]}
    _log(state, "schema_analyzer", f"✅ {len(result.get('columns',[]))} columns classified")
    return {"schema": result}


def node_rule_interpreter(state: ETLState) -> dict:
    """Agent 2 — LCEL chain: prompt | llm | JsonOutputParser (batched)."""
    rules  = state["raw_rules"]
    _log(state, "rule_interpreter", f"Interpreting {len(rules)} rules...")
    llm    = _get_llm(state)
    chain  = _make_rule_chain(llm)
    all_out = []
    for i in range(0, len(rules), 30):
        batch  = rules[i:i+30]
        result = chain.invoke({
            "input_columns": str(state["input_columns"]),
            "rules_json":    json.dumps(batch, indent=2),
        })
        if isinstance(result, list) and result:
            all_out.extend(result)
        else:
            all_out.extend(_regex_interpret(batch, state["input_columns"]))

    # Align output_col
    for idx, rule in enumerate(all_out):
        if not rule.get("output_col") and idx < len(rules):
            rule["output_col"] = rules[idx].get("output_col","")

    _log(state, "rule_interpreter", f"✅ {len(all_out)} rules interpreted")
    return {"interpreted_rules": all_out}


def node_precision_guard(state: ETLState) -> dict:
    """Agent 3 — Deterministic: tag arithmetic ops for BigDecimal engine."""
    rules  = state["interpreted_rules"]
    arith  = {"multiply","subtract","divide","add_constant"}
    issues, cnt = [], 0
    for rule in rules:
        op = rule.get("operation","")
        if op in arith:
            rule.update({"precision":4,"decimal_engine":True,"float_prohibited":True})
            cnt += 1
            if op != "add_constant" and not rule.get("operand_b"):
                issues.append(f"⚠ {rule['output_col']}: missing operand_b")
    _log(state, "precision_guard", f"✅ {cnt} ops → Decimal engine, {len(issues)} warnings")
    return {"interpreted_rules": rules, "errors": state.get("errors",[]) + issues}


def node_db_mapping(state: ETLState) -> dict:
    """Agent 4 — Deterministic: resolve DB lookups for all rows × lookup fields."""
    import re
    rules    = state["interpreted_rules"]
    in_rows  = state["input_rows"]
    db_rules = [r for r in rules if r.get("operation")=="db_lookup"]
    _log(state, "db_mapping", f"Resolving {len(db_rules)} DB fields × {len(in_rows)} rows...")
    resolved = {}
    for ri, row in enumerate(in_rows):
        isin = ""
        for cand in ["ISIN","isin"]:
            if cand in row and row[cand]: isin=str(row[cand]).strip(); break
        if not isin:
            for v in row.values():
                if re.match(r"INF\w+",str(v)): isin=str(v).strip(); break
        for rule in db_rules:
            resolved[(ri, rule["output_col"])] = db_lookup(isin, rule.get("db_key",""))
    found = sum(1 for v in resolved.values() if v!="N/A")
    _log(state, "db_mapping", f"✅ {found}/{len(resolved)} resolved")
    # Serialise keys for state storage (tuples → "ri:col")
    return {"db_resolved": {f"{k[0]}:{k[1]}": v for k,v in resolved.items()}}


def node_output_mapping(state: ETLState) -> dict:
    """Agent 5 — Deterministic execution engine."""
    rules     = state["interpreted_rules"]
    in_rows   = state["input_rows"]
    out_cols  = state["output_columns"]
    db_serial = state.get("db_resolved", {})
    # Deserialise db_resolved
    db_resolved = {}
    for k, v in db_serial.items():
        ri, col = k.split(":", 1)
        db_resolved[(int(ri), col)] = v

    _log(state, "output_mapping", f"Executing {len(rules)} rules × {len(in_rows)} rows...")
    output_rows = []
    for ri, row in enumerate(in_rows):
        out = _execute_row(rules, row, ri, db_resolved)
        for col in out_cols: out.setdefault(col,"")
        output_rows.append(out)
    _log(state, "output_mapping", f"✅ {len(output_rows)} rows × {len(out_cols)} cols")
    return {"output_rows": output_rows}


def node_drift_detection(state: ETLState) -> dict:
    """Agent 6 — LCEL chain: prompt | llm | JsonOutputParser."""
    _log(state, "drift_detection", "Running drift detection...")
    llm     = _get_llm(state)
    chain   = _make_drift_chain(llm)
    in_rows = state["input_rows"]
    out_rows= state["output_rows"]
    rules   = state["interpreted_rules"]
    samples = [{"row":i+1,"input":inp,"output_sample":{k:out[k] for k in list(out)[:12]}}
               for i,(inp,out) in enumerate(zip(in_rows[:3],out_rows[:3]))]
    result  = chain.invoke({
        "rules_sample": json.dumps(rules[:15],indent=2),
        "io_samples":   json.dumps(samples,indent=2),
    })
    if not result: result={"drift_detected":False,"issues":[],"summary":"No drift (fallback)."}
    drift = result.get("drift_detected",False)
    _log(state, "drift_detection", f"✅ drift={'YES ⚠' if drift else 'none'}")
    return {"drift_result": result, "drift_detected": drift}


def node_rule_optimizer(state: ETLState) -> dict:
    """Agent 7 — LCEL chain: prompt | llm | JsonOutputParser."""
    rules = state["interpreted_rules"]
    _log(state, "rule_optimizer", f"Optimizing {len(rules)} rules...")
    llm   = _get_llm(state)
    chain = _make_opt_chain(llm)
    dist  = {}
    for r in rules: dist[r.get("operation","?")] = dist.get(r.get("operation","?"),0)+1
    result = chain.invoke({"op_dist": json.dumps(dist), "rules_sample": json.dumps(rules[:10],indent=2)})
    if not result:
        db_f = [r["output_col"] for r in rules if r.get("operation")=="db_lookup"]
        result = {"optimizations":[
            {"type":"batch_db","description":f"{len(db_f)} DB lookups → batch per ISIN","impact":"high","affected_fields":db_f},
            {"type":"template","description":"Repeating arithmetic groups → template","impact":"medium","affected_fields":[]},
        ],"summary":str(dist)}
    _log(state, "rule_optimizer", f"✅ {len(result.get('optimizations',[]))} suggestions")
    return {"opt_result": result}


def node_test_generator(state: ETLState) -> dict:
    """Agent 8 — LCEL chain: generate + deterministic run."""
    rules   = state["interpreted_rules"]
    in_rows = state["input_rows"]
    out_rows= state["output_rows"]
    _log(state, "test_generator", f"Generating test cases for {len(rules)} rules...")
    llm   = _get_llm(state)
    chain = _make_test_chain(llm)
    tcs   = chain.invoke({
        "input_columns": str(state["input_columns"]),
        "rules_sample":  json.dumps(rules[:20],indent=2),
        "sample_row":    json.dumps(in_rows[0] if in_rows else {},default=str),
    })
    if not tcs or not isinstance(tcs,list):
        tcs = [{"rule_field":r["output_col"],"operation":r.get("operation",""),
                "test_name":f"happy_{r['output_col']}","inputs":in_rows[0] if in_rows else {},
                "expected_formula":r.get("operation",""),"edge_case":False}
               for r in rules[:12]]

    # Run tests deterministically
    arith = {"multiply","subtract","divide","add_constant"}
    passed=failed=0; details=[]
    for tc in tcs:
        field,op = tc.get("rule_field",""),tc.get("operation","")
        for ri,row in enumerate(out_rows):
            actual = row.get(field,None)
            if actual is None:
                failed+=1; details.append({"test":tc["test_name"],"row":ri+1,"status":"FAIL","reason":f"'{field}' missing"})
                continue
            if op in arith:
                try: Decimal(str(actual)); passed+=1; details.append({"test":tc["test_name"],"row":ri+1,"status":"PASS","actual":str(actual)})
                except: failed+=1; details.append({"test":tc["test_name"],"row":ri+1,"status":"FAIL","reason":f"'{actual}' not decimal"})
            else:
                passed+=1; details.append({"test":tc["test_name"],"row":ri+1,"status":"PASS","actual":str(actual)})

    _log(state,"test_generator",f"✅ {len(tcs)} cases, {passed}/{passed+failed} passed")
    return {"test_cases":tcs,"test_results":{"passed":passed,"failed":failed,"total":passed+failed,"details":details[:40]}}


def node_anomaly_detector(state: ETLState) -> dict:
    """Agent 9 — LCEL chain: prompt | llm | JsonOutputParser."""
    out_rows = state["output_rows"]
    _log(state, "anomaly_detector", f"Scanning {len(out_rows)} rows...")
    stats: dict[str,list] = {}
    for row in out_rows:
        for k,v in row.items():
            if str(v) in ("N/A","","None"): continue
            try: stats.setdefault(k,[]).append(float(Decimal(str(v))))
            except: pass
    summary = {k:{"min":min(v),"max":max(v),"mean":round(sum(v)/len(v),6),"zeros":v.count(0.0),"count":len(v)}
               for k,v in stats.items() if v}
    llm   = _get_llm(state)
    chain = _make_anomaly_chain(llm)
    result= chain.invoke({"stats":json.dumps(summary,indent=2),"row_count":str(len(out_rows))})
    if not result:
        anoms=[{"field":k,"type":"all_zeros","detail":"All 0","severity":"high"}
               for k,s in summary.items() if s["zeros"]==s["count"]>0]
        result={"anomalies":anoms,"clean_fields":len(summary)-len(anoms),"summary":f"{len(anoms)} anomalies"}
    _log(state,"anomaly_detector",f"✅ {len(result.get('anomalies',[]))} anomalies")
    return {"anomaly_result": result}


def node_excel_writer(state: ETLState) -> dict:
    """Write the output Excel. Deterministic."""
    _log(state, "excel_writer", f"Writing {len(state['output_rows'])} rows × {len(state['output_columns'])} cols...")
    _write_excel(state, None)
    _log(state, "excel_writer", f"✅ Written: {state['out_path']}")
    return {}


def node_runtime_validator(state: ETLState) -> dict:
    """Agent 10 — compute expected + LCEL spec chain + validate Excel cell-by-cell."""
    import openpyxl
    rules    = state["interpreted_rules"]
    in_rows  = state["input_rows"]
    out_cols = state["output_columns"]
    out_path = state["out_path"]
    db_serial= state.get("db_resolved",{})
    db_resolved = {(int(k.split(":")[0]), k.split(":",1)[1]):v for k,v in db_serial.items()}

    _log(state, "runtime_validator", "Computing expected values...")
    expected_rows = []
    for ri, row in enumerate(in_rows):
        exp = _execute_row(rules, row, ri, db_resolved)
        expected_rows.append(exp)

    # LLM-generated specs via LCEL
    llm   = _get_llm(state)
    chain = _make_spec_chain(llm)
    sample_exp = {k:v for k,v in expected_rows[0].items() if k in out_cols[:15]} if expected_rows else {}
    specs_raw = chain.invoke({
        "rules_sample": json.dumps(rules[:25],indent=2),
        "sample_exp":   json.dumps(sample_exp),
    })
    specs_raw = specs_raw if isinstance(specs_raw,list) else []
    spec_map  = {s["field"]:s for s in specs_raw}

    # Fallback specs
    spec_labels = {"copy":"Copy {a} from input","multiply":"{a} × {b} (4dp)",
                   "subtract":"{a} − {b} (4dp)","divide":"{a} ÷ {b} (4dp)",
                   "add_constant":"{a} + {c} (4dp)","db_lookup":"DB(ISIN→{k})"}
    for rule in rules:
        col = rule.get("output_col","")
        if col not in spec_map:
            op=rule.get("operation","")
            tmpl=spec_labels.get(op,op)
            spec=tmpl.format(a=rule.get("operand_a","?"),b=rule.get("operand_b","?"),
                             c=rule.get("constant","?"),k=rule.get("db_key","?"))
            spec_map[col]={"field":col,"spec":spec,"category":op if op in("copy","db_lookup") else "arithmetic"}

    # Validate Excel
    _log(state, "runtime_validator", f"Validating {out_path} cell-by-cell...")
    try:
        wb = openpyxl.load_workbook(out_path, data_only=True)
        ws = wb.active
    except Exception as e:
        return {"validation":{"status":"ERROR","error":str(e),"passed":0,"failed":0,"total":0},"validation_pass":False}

    headers = [ws.cell(row=1,column=c).value for c in range(1,ws.max_column+1)]
    col_idx = {h:i+1 for i,h in enumerate(headers) if h}
    rule_map= {r.get("output_col",""):r for r in rules}

    results=[]; passed=failed=skipped=0
    for ri2, exp_row in enumerate(expected_rows):
        excel_row = ri2+2
        for col in out_cols:
            if col not in col_idx: skipped+=1; continue
            raw    = ws.cell(row=excel_row,column=col_idx[col]).value
            actual = str(raw).strip() if raw is not None else ""
            exp    = str(exp_row.get(col,"")).strip()
            rule   = rule_map.get(col,{})
            op     = rule.get("operation","")
            spec   = spec_map.get(col,{}).get("spec","")
            status,reason = _compare_cell(actual,exp,op)
            if status=="PASS": passed+=1
            else:              failed+=1
            results.append({"row":ri2+1,"field":col,"operation":op,"spec":spec,
                             "expected":exp,"actual":actual,"status":status,"reason":reason})

    total = passed+failed
    pct   = round(100*passed/total,1) if total else 0.0
    field_fails = {}
    for r in results:
        if r["status"]=="FAIL": field_fails[r["field"]]=field_fails.get(r["field"],0)+1

    validation = {
        "status":"PASS" if failed==0 else "FAIL",
        "passed":passed,"failed":failed,"skipped":skipped,
        "total":total,"pass_pct":pct,"field_failures":field_fails,
        "details":results,"failures_only":[r for r in results if r["status"]=="FAIL"],
    }
    val_pass = failed == 0
    _log(state,"runtime_validator",f"✅ {passed}/{total} cells passed ({pct}%)")
    return {"expected_rows":expected_rows,"spec_map":spec_map,
            "validation":validation,"validation_pass":val_pass}


def node_excel_rewriter(state: ETLState) -> dict:
    """Rewrite Excel with validation highlights."""
    _log(state,"excel_rewriter","Rewriting Excel with validation highlights...")
    _write_excel(state, state.get("validation"))
    _log(state,"excel_rewriter",f"✅ Final Excel: {state['out_path']}")
    return {}


def node_final_report(state: ETLState) -> dict:
    """Print pipeline summary."""
    tv  = state.get("validation",{})
    tr  = state.get("test_results",{})
    _log(state,"final_report","Pipeline complete ✅")
    print(f"\n{'═'*65}")
    print(f"  PIPELINE COMPLETE — LangGraph ETL")
    print(f"{'═'*65}")
    print(f"  Input rows        : {len(state.get('input_rows',[]))}")
    print(f"  Output cols       : {len(state.get('output_columns',[]))}")
    print(f"  Rules executed    : {len(state.get('interpreted_rules',[]))}")
    print(f"  Drift detected    : {state.get('drift_detected',False)}")
    print(f"  Anomalies         : {len(state.get('anomaly_result',{}).get('anomalies',[]))}")
    print(f"  In-memory tests   : {tr.get('passed',0)}/{tr.get('total',0)} passed")
    print(f"  Excel validation  : {tv.get('status','?')} — {tv.get('passed',0)}/{tv.get('total',0)} cells ({tv.get('pass_pct',0):.1f}%)")
    print(f"  Output file       : {state.get('out_path','')}")
    print(f"  Retry count       : {state.get('retry_count',0)}")
    print(f"{'═'*65}")
    return {}


# ══════════════════════════════════════════════════════════════════════════════
# CONDITIONAL ROUTERS  (LangGraph routing functions)
# ══════════════════════════════════════════════════════════════════════════════

def route_after_drift(state: ETLState) -> str:
    """After drift detection: always continue (drift is logged, not blocking)."""
    return "rule_optimizer"


def route_after_validation(state: ETLState) -> str:
    """
    After runtime validation:
    - PASS → rewrite Excel with highlights, finish
    - FAIL + retries left → re-run optimizer and recheck
    - FAIL + no retries → accept and finish
    """
    if state.get("validation_pass", False):
        return "excel_rewriter"
    retry = state.get("retry_count", 0)
    max_r = state.get("max_retries", 2)
    if retry < max_r:
        return "retry"
    return "excel_rewriter"    # accept after max retries


def node_retry(state: ETLState) -> dict:
    """Increment retry counter — graph loops back to rule_optimizer."""
    rc = state.get("retry_count", 0) + 1
    _log(state, "retry", f"Retry {rc}/{state.get('max_retries',2)} — re-running optimizer...")
    return {"retry_count": rc}


# ══════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def _get_llm(state: ETLState):
    return create_llm(
        provider   = state.get("provider","gemini"),
        model      = state.get("model") or "",
        temperature= 0.0,
        max_tokens = 4096,
    )


def _log(state: ETLState, agent: str, msg: str):
    entry = {"agent": agent, "msg": msg, "ts": datetime.now(timezone.utc).isoformat()}
    print(f"[{agent}] {msg}")
    # State is a dict — append to log list
    if "agent_log" in state and isinstance(state["agent_log"], list):
        state["agent_log"].append(entry)


def _write_excel(state: ETLState, validation):
    out_rows  = state["output_rows"]
    out_cols  = state["output_columns"]
    out_path  = state["out_path"]
    rules     = state.get("interpreted_rules",[])
    tv        = validation or {}

    wb   = Workbook()
    hf   = Font(name="Arial",bold=True,color="FFFFFF",size=10)
    hfill= PatternFill("solid",fgColor="1F4E79")
    cfill= PatternFill("solid",fgColor="E8F4FD")
    dfill= PatternFill("solid",fgColor="E8F5E9")
    pfill= PatternFill("solid",fgColor="FFF9C4")
    ffill= PatternFill("solid",fgColor="FFCCCC")
    gfill= PatternFill("solid",fgColor="CCFFCC")
    thin = Side(style="thin",color="CCCCCC")
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)

    copy_cols = {r["output_col"] for r in rules if r.get("operation")=="copy"}
    db_cols   = {r["output_col"] for r in rules if r.get("operation")=="db_lookup"}
    calc_cols = {r["output_col"] for r in rules if r.get("operation") in("multiply","subtract","divide","add_constant")}
    failed_cells = {(d["row"],d["field"]) for d in tv.get("failures_only",[])}

    # ── Output sheet ──────────────────────────────────────────────────────────
    ws=wb.active; ws.title="Output_Data"
    for ci,col in enumerate(out_cols,1):
        c=ws.cell(row=1,column=ci,value=col)
        c.font=hf; c.fill=hfill; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=bdr; ws.column_dimensions[get_column_letter(ci)].width=max(14,len(col)+2)
    ws.row_dimensions[1].height=40; ws.freeze_panes="A2"

    for ri,rd in enumerate(out_rows,2):
        for ci,col in enumerate(out_cols,1):
            val=rd.get(col,"")
            c=ws.cell(row=ri,column=ci,value=val)
            c.border=bdr; c.alignment=Alignment(horizontal="left",vertical="center")
            c.font=Font(name="Arial",size=9)
            rn=ri-1
            if (rn,col) in failed_cells:  c.fill=ffill
            elif tv.get("passed") and (rn,col) not in failed_cells and col in (calc_cols|db_cols|copy_cols): c.fill=gfill
            elif col in copy_cols:  c.fill=pfill
            elif col in db_cols:    c.fill=dfill
            elif col in calc_cols:  c.fill=cfill
        ws.row_dimensions[ri].height=18

    # ── Validation Report ─────────────────────────────────────────────────────
    wv=wb.create_sheet("Validation_Report")
    wv["A1"]="Agent 10 — Runtime Validation Report (LangGraph)"
    wv["A1"].font=Font(bold=True,size=13,color="1F4E79"); wv.merge_cells("A1:G1")
    if tv:
        for i,(lbl,val) in enumerate([
            ("Status",tv.get("status","?")),("Total cells",str(tv.get("total",0))),
            ("Passed",str(tv.get("passed",0))),("Failed",str(tv.get("failed",0))),
            ("Pass %",f"{tv.get('pass_pct',0):.1f}%"),("Retry count",str(state.get("retry_count",0))),
        ],3):
            wv.cell(row=i,column=1,value=lbl).font=Font(bold=True,name="Arial",size=10)
            wv.cell(row=i,column=2,value=val).font=Font(name="Arial",size=10)
        hdrs=["Row","Field","Operation","Spec","Expected","Actual","Status","Reason"]
        for ci2,h in enumerate(hdrs,1):
            c=wv.cell(row=12,column=ci2,value=h); c.font=hf; c.fill=hfill
        for ri2,d in enumerate(tv.get("details",[])[:500],13):
            vs=[d["row"],d["field"],d["operation"],d["spec"],d["expected"],d["actual"],d["status"],d.get("reason","")]
            for ci2,v in enumerate(vs,1):
                c=wv.cell(row=ri2,column=ci2,value=v)
                c.font=Font(name="Arial",size=8); c.border=bdr
                c.fill=ffill if d["status"]=="FAIL" else gfill
        for ci2,w in enumerate([6,20,14,40,16,16,8,30],1):
            wv.column_dimensions[get_column_letter(ci2)].width=w

    # ── Execution Graph ───────────────────────────────────────────────────────
    wg=wb.create_sheet("LangGraph_Flow")
    wg["A1"]="LangGraph Execution Flow (Mermaid)"; wg["A1"].font=Font(bold=True,size=12)
    mermaid = "\n".join([
        "graph TD",
        "    START((START)) --> loader",
        "    loader --> schema_analyzer",
        "    schema_analyzer --> rule_interpreter",
        "    rule_interpreter --> precision_guard",
        "    precision_guard --> db_mapping",
        "    db_mapping --> output_mapping",
        "    output_mapping --> drift_detection",
        "    drift_detection -->|route| rule_optimizer",
        "    rule_optimizer --> test_generator",
        "    test_generator --> anomaly_detector",
        "    anomaly_detector --> excel_writer",
        "    excel_writer --> runtime_validator",
        "    runtime_validator -->|PASS| excel_rewriter",
        "    runtime_validator -->|FAIL+retry| retry",
        "    retry --> rule_optimizer",
        "    excel_rewriter --> final_report",
        "    final_report --> END((END))",
    ])
    for i,line in enumerate(mermaid.split("\n"),3):
        wg.cell(row=i,column=1,value=line).font=Font(name="Courier New",size=9)
    wg.column_dimensions["A"].width=60

    # ── Audit ─────────────────────────────────────────────────────────────────
    wa=wb.create_sheet("Audit_Report")
    wa["A1"]="LangGraph ETL Audit"; wa["A1"].font=Font(bold=True,size=14,color="1F4E79")
    wa.merge_cells("A1:D1")
    arows=[
        ("Timestamp",     state.get("timestamp","")),
        ("LLM Provider",  state.get("provider","")),
        ("LLM Model",     state.get("model","")),
        ("Input file",    state.get("input_file","")),
        ("Rule file",     state.get("rule_file","")),
        ("Total rules",   str(len(state.get("raw_rules",[])))),
        ("Input rows",    str(len(state.get("input_rows",[])))),
        ("Output cols",   str(len(state.get("output_columns",[])))),
        ("Drift",         str(state.get("drift_detected",False))),
        ("Anomalies",     str(len(state.get("anomaly_result",{}).get("anomalies",[])))),
        ("In-mem passed", str(state.get("test_results",{}).get("passed",0))),
        ("Excel passed",  str(tv.get("passed",0))),
        ("Excel failed",  str(tv.get("failed",0))),
        ("Pass %",        f"{tv.get('pass_pct',0):.1f}%"),
        ("Retries",       str(state.get("retry_count",0))),
        ("Optimizations", state.get("opt_result",{}).get("summary","")),
    ]
    for i,(lbl,val) in enumerate(arows,3):
        wa.cell(row=i,column=1,value=lbl).font=Font(bold=True,name="Arial",size=10)
        c=wa.cell(row=i,column=2,value=val); c.font=Font(name="Arial",size=10)
        c.alignment=Alignment(wrap_text=True)
    wa.column_dimensions["A"].width=22; wa.column_dimensions["B"].width=65

    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
# BUILD + COMPILE THE LANGGRAPH
# ══════════════════════════════════════════════════════════════════════════════

def build_graph(checkpointer=None) -> "CompiledGraph":
    """
    Wire all 10 agent nodes into a StateGraph with:
    - Linear edges for most steps
    - Conditional edge after drift_detection
    - Conditional edge after runtime_validator (retry loop)
    """
    graph = StateGraph(ETLState)

    # ── Register nodes ────────────────────────────────────────────────────────
    graph.add_node("loader",           node_loader)
    graph.add_node("schema_analyzer",  node_schema_analyzer)
    graph.add_node("rule_interpreter", node_rule_interpreter)
    graph.add_node("precision_guard",  node_precision_guard)
    graph.add_node("db_mapping",       node_db_mapping)
    graph.add_node("output_mapping",   node_output_mapping)
    graph.add_node("drift_detection",  node_drift_detection)
    graph.add_node("rule_optimizer",   node_rule_optimizer)
    graph.add_node("test_generator",   node_test_generator)
    graph.add_node("anomaly_detector", node_anomaly_detector)
    graph.add_node("excel_writer",     node_excel_writer)
    graph.add_node("runtime_validator",node_runtime_validator)
    graph.add_node("excel_rewriter",   node_excel_rewriter)
    graph.add_node("retry",            node_retry)
    graph.add_node("final_report",     node_final_report)

    # ── Entry point ───────────────────────────────────────────────────────────
    graph.set_entry_point("loader")

    # ── Linear edges ─────────────────────────────────────────────────────────
    graph.add_edge("loader",           "schema_analyzer")
    graph.add_edge("schema_analyzer",  "rule_interpreter")
    graph.add_edge("rule_interpreter", "precision_guard")
    graph.add_edge("precision_guard",  "db_mapping")
    graph.add_edge("db_mapping",       "output_mapping")
    graph.add_edge("output_mapping",   "drift_detection")
    graph.add_edge("rule_optimizer",   "test_generator")
    graph.add_edge("test_generator",   "anomaly_detector")
    graph.add_edge("anomaly_detector", "excel_writer")
    graph.add_edge("excel_writer",     "runtime_validator")
    graph.add_edge("excel_rewriter",   "final_report")
    graph.add_edge("final_report",     END)

    # ── Conditional: drift_detection → rule_optimizer (always, via router) ───
    graph.add_conditional_edges(
        "drift_detection",
        route_after_drift,
        {"rule_optimizer": "rule_optimizer"},
    )

    # ── Conditional: runtime_validator → excel_rewriter | retry ──────────────
    graph.add_conditional_edges(
        "runtime_validator",
        route_after_validation,
        {
            "excel_rewriter": "excel_rewriter",
            "retry":          "retry",
        },
    )

    # ── Retry loop: retry → rule_optimizer ───────────────────────────────────
    graph.add_edge("retry", "rule_optimizer")

    return graph.compile(checkpointer=checkpointer)


# ══════════════════════════════════════════════════════════════════════════════
# RUN
# ══════════════════════════════════════════════════════════════════════════════

def run(
    input_file: str,
    rule_file:  str,
    out_path:   str,
    provider:   str  = "gemini",
    model:      str  = "gemini-2.5-flash",
    stream:     bool = False,
    thread_id:  str  = "etl-run-1",
) -> dict:
    print(f"\n{'═'*65}")
    print(f"  LANGGRAPH INVESTMENT ETL — 10 AGENTS")
    print(f"  Provider: {provider}  |  Model: {model or 'default'}")
    print(f"{'═'*65}")

    checkpointer = MemorySaver()
    app          = build_graph(checkpointer=checkpointer)
    config       = {"configurable": {"thread_id": thread_id}}

    state0 = initial_state(input_file, rule_file, out_path, provider, model)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    if stream:
        print("\n[LangGraph] Streaming mode — printing after each node:\n")
        final_state = state0
        for event in app.stream(state0, config=config):
            node_name = list(event.keys())[0]
            print(f"  ─── completed: {node_name}")
            final_state = event[node_name]
        return final_state
    else:
        return app.invoke(state0, config=config)


# ── CLI ───────────────────────────────────────────────────────────────────────

def _cli():
    p = argparse.ArgumentParser(description="LangGraph Investment ETL — 10 Agents")
    p.add_argument("--input",    required=True)
    p.add_argument("--rules",    required=True)
    p.add_argument("--out",      default="output/Investment_LangGraph.xlsx")
    p.add_argument("--provider", default="claude", choices=["claude","gemini","ollama"])
    p.add_argument("--model",    default="gemini-2.5-flash")
    p.add_argument("--stream",   action="store_true", help="Use streaming mode")
    p.add_argument("--thread-id",default="etl-run-1")
    args = p.parse_args()
    run(input_file=args.input, rule_file=args.rules, out_path=args.out,
        provider=args.provider, model=args.model, stream=args.stream,
        thread_id=args.thread_id)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        _cli()
    else:
        run(
            input_file = "/mnt/user-data/uploads/Investment_Sample_Data__1_.xlsx",
            rule_file  = "/mnt/user-data/uploads/Rule_Mapping_92_English.xlsx",
            out_path   = "/mnt/user-data/outputs/Investment_LangGraph.xlsx",
            provider   = "claude",
            stream     = False,
        )
